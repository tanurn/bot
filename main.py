import json
import os
from typing import Dict, List
import re
import logging
import asyncio
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Union

from aiogram import Bot, Dispatcher, types, F, Router
from aiogram.types import Message, FSInputFile, ReplyKeyboardRemove
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder, InlineKeyboardMarkup
from aiogram.client.default import DefaultBotProperties
from telethon import TelegramClient
from telethon.tl.types import Message as TgMessage
from telethon.tl.patched import MessageService
from telethon.errors import FloodWaitError, ChannelPrivateError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# Загрузка конфигурации из переменных окружения
API_ID = int(os.getenv("TELEGRAM_API_ID", "29452772"))
API_HASH = os.getenv("TELEGRAM_API_HASH", "bf975d37f0f1a82f0abbf02170f8bb9d")
BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "8087512664:AAEabKVUKZpjO4DvPxa37TN8YJbIhslwCCg")
PHONE_NUMBER = os.getenv("TELEGRAM_PHONE_NUMBER", '+79872498759')  # Номер телефона пользователя
SESSION_NAME = "user_session"
ALLOWED_USER_IDS = [int(x) for x in os.getenv("ALLOWED_USER_IDS", "1288093529,1288093529").split(",")]
SUPPLIER_CHANNELS = [int(channel_id.strip()) for channel_id in os.getenv("SUPPLIER_CHANNELS", "-1002470102452,-1001785205512").split(",") if channel_id.strip()]
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO")
LOG_FILE = os.getenv("LOG_FILE", "bot_operations.log")

# Настройка правил ценообразования
PRICE_RULES = {
    "iPhone": {"base_margin": 0.15, "premium_bonus": 0.1},
    "Apple Watch": {"fixed_margin": 5000},
    "MacBook": {"base_margin": 0.12},
    "iPad": {"base_margin": 0.13},
    "AirPods": {"base_margin": 0.14},
    "iMac": {"base_margin": 0.11},
    "Samsung": {"base_margin": 0.08},
    "Xiaomi": {"base_margin": 0.07},
    "Аксессуары": {"base_margin": 0.20},
    "default": {"margin": 0.1}
}

# Глобальное хранилище данных
DATA_STORAGE = {
    "telethon_client": None,
    "channel_names": {},  # Кэш названий каналов
    "users": {}           # Данные по пользователям {user_id: {last_channel, last_message_ids, processed_data}}
}

# Настройка логирования
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Инициализация бота и диспетчера
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
dp = Dispatcher()
router = Router()
dp.include_router(router)


# Глобальные переменные для хранения настроек
PARSING_SETTINGS = {
    "remove_emojis": True,
    "replace_flags": True,
    "process_sim_types": True,
    "delivery_icons": {"♻️": "Рефабрикат", "⌛️": "Ожидается"}
}

PRICE_SETTINGS = PRICE_RULES  # Используем уже существующие правила

DATA_SOURCES = SUPPLIER_CHANNELS  # Используем уже существующие каналы

# Файлы для сохранения настроек
SETTINGS_DIR = "settings"
os.makedirs(SETTINGS_DIR, exist_ok=True)
PARSING_SETTINGS_FILE = os.path.join(SETTINGS_DIR, "parsing_settings.json")
PRICE_SETTINGS_FILE = os.path.join(SETTINGS_DIR, "price_settings.json")
DATA_SOURCES_FILE = os.path.join(SETTINGS_DIR, "data_sources.json")

# Загрузка сохраненных настроек при запуске
def load_settings():
    global PARSING_SETTINGS, PRICE_SETTINGS, DATA_SOURCES
    
    try:
        if os.path.exists(PARSING_SETTINGS_FILE):
            with open(PARSING_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                PARSING_SETTINGS = json.load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки настроек парсинга: {str(e)}")
    
    try:
        if os.path.exists(PRICE_SETTINGS_FILE):
            with open(PRICE_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                PRICE_SETTINGS = json.load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки правил ценообразования: {str(e)}")
    
    try:
        if os.path.exists(DATA_SOURCES_FILE):
            with open(DATA_SOURCES_FILE, 'r', encoding='utf-8') as f:
                DATA_SOURCES = json.load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки источников данных: {str(e)}")

# Сохранение настроек
def save_parsing_settings():
    with open(PARSING_SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(PARSING_SETTINGS, f, ensure_ascii=False, indent=2)

def save_price_settings():
    with open(PRICE_SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(PRICE_SETTINGS, f, ensure_ascii=False, indent=2)

def save_data_sources():
    with open(DATA_SOURCES_FILE, 'w', encoding='utf-8') as f:
        json.dump(DATA_SOURCES, f, ensure_ascii=False, indent=2)

# Загрузка настроек при старте
load_settings()


# Состояния FSM
class Form(StatesGroup):
    waiting_for_message_ids = State()
    waiting_for_channel_selection = State()
    waiting_for_verification_code = State()
    waiting_for_parsing_settings = State()
    waiting_for_price_category = State()
    waiting_for_price_value = State()
    waiting_for_data_sources = State()
    waiting_for_new_channel = State()

# Уровни меню для навигации
class MenuLevel:
    MAIN = "main"
    DATA_PROCESSING = "data_processing"
    SETTINGS = "settings"
    CHANNELS = "channels"

# Middleware для контроля доступа
class AccessMiddleware:
    async def __call__(self, handler, event: types.Message, data):
        if event.from_user.id not in ALLOWED_USER_IDS:
            await event.answer("🚫 Доступ запрещен! Ваш ID не в списке разрешенных.")
            logger.warning(f"Unauthorized access attempt: {event.from_user.id}")
            return
        return await handler(event, data)

# Middleware для защиты от спама
class ThrottlingMiddleware:
    def __init__(self, delay=1.0):
        self.delay = delay
        self.last_processed = {}

    async def __call__(self, handler, event: types.Message, data):
        user_id = event.from_user.id
        current_time = datetime.now().timestamp()
        
        if user_id in self.last_processed:
            elapsed = current_time - self.last_processed[user_id]
            if elapsed < self.delay:
                await event.answer("⚠️ Слишком много запросов. Пожалуйста, подождите.")
                return
                
        self.last_processed[user_id] = current_time
        return await handler(event, data)

# Регистрация middleware
router.message.middleware(AccessMiddleware())
router.message.middleware(ThrottlingMiddleware())

# =============================
# ИНИЦИАЛИЗАЦИЯ TELETHON КЛИЕНТА
# =============================

async def init_telethon_client():
    """Инициализация клиента Telethon с пользовательским аккаунтом"""
    logger.info("Инициализация Telethon клиента...")
    
    client = TelegramClient(SESSION_NAME, API_ID, API_HASH)
    await client.start(phone=lambda: PHONE_NUMBER)
    
    if not await client.is_user_authorized():
        # Если требуется код подтверждения
        await request_verification_code()
        return None
    
    logger.info("Telethon клиент успешно авторизован")
    
    # Кэшируем названия каналов
    await refresh_channel_names()
    
    DATA_STORAGE["telethon_client"] = client
    return client

async def shutdown_telethon_client():
    """Завершение работы Telethon клиента"""
    client = DATA_STORAGE.get("telethon_client")
    if client:
        logger.info("Завершение работы Telethon клиента...")
        await client.disconnect()
        logger.info("Telethon клиент успешно остановлен")

async def request_verification_code():
    """Запрос кода подтверждения у пользователя"""
    for user_id in ALLOWED_USER_IDS:
        await bot.send_message(
            user_id,
            "🔑 Требуется код подтверждения для авторизации в Telegram. "
            "Пожалуйста, введите код, который вы получили в Telegram, используя команду /code <ваш_код>"
        )

async def refresh_channel_names():
    """Обновление кэша названий каналов"""
    client = DATA_STORAGE.get("telethon_client")
    if not client:
        return
    
    for channel_id in SUPPLIER_CHANNELS:
        try:
            entity = await client.get_entity(channel_id)
            DATA_STORAGE["channel_names"][channel_id] = entity.title
        except Exception as e:
            logger.error(f"Ошибка получения названия канала {channel_id}: {str(e)}")
            DATA_STORAGE["channel_names"][channel_id] = f"Канал {channel_id}"

def get_channel_name(channel_id: int) -> str:
    """Получение названия канала из кэша"""
    return DATA_STORAGE["channel_names"].get(channel_id, f"Канал {channel_id}")

async def fetch_message_from_chat(channel_id: Any, message_id: int) -> Optional[str]:
    """Получение сообщения из канала по ID"""
    client = DATA_STORAGE.get("telethon_client")
    if not client:
        logger.error("Telethon клиент не инициализирован")
        return None
    
    try:
        # Пытаемся получить сообщение
        message = await client.get_messages(channel_id, ids=message_id)
        
        # Проверяем, что это текстовое сообщение
        if not message or isinstance(message, MessageService):
            logger.warning(f"Сообщение {message_id} не найдено или не является текстовым")
            return None
            
        # Возвращаем текст сообщения
        return message.text
        
    except FloodWaitError as e:
        logger.error(f"Ошибка FloodWait: необходимо подождать {e.seconds} секунд")
        await asyncio.sleep(e.seconds + 2)
        return await fetch_message_from_chat(channel_id, message_id)
        
    except ChannelPrivateError:
        logger.error(f"Нет доступа к каналу {channel_id}")
        return None
        
    except Exception as e:
        logger.error(f"Ошибка при получении сообщения {message_id}: {str(e)}")
        return None

async def fetch_messages(channel_id: Any, message_ids: List[int]) -> Dict[int, str]:
    """Получение нескольких сообщений из канала"""
    results = {}
    for msg_id in message_ids:
        text = await fetch_message_from_chat(channel_id, msg_id)
        if text:
            results[msg_id] = text
        # Задержка для избежания FloodWait
        await asyncio.sleep(0.5)
    return results

# =============================
# КОНСТРУКТОРЫ КЛАВИАТУР
# =============================

def build_main_menu() -> InlineKeyboardMarkup:
    """Главное меню бота"""
    builder = InlineKeyboardBuilder()
    builder.button(text="📥 Загрузить сообщения", callback_data="get_messages")
    builder.button(text="🔄 Обработать данные", callback_data="process_data")
    builder.button(text="📋 Список каналов", callback_data="list_channels")
    builder.button(text="⚙️ Настройки", callback_data="settings")
    builder.button(text="📤 Экспорт в XLSX", callback_data="export_xlsx")
    builder.adjust(2, 2, 1)
    return builder.as_markup()

def build_back_button(back_to: str) -> InlineKeyboardMarkup:
    """Кнопка 'Назад' для возврата в предыдущее меню"""
    builder = InlineKeyboardBuilder()
    builder.button(text="🔙 Назад", callback_data=f"back_to_{back_to}")
    return builder.as_markup()

def build_channels_menu() -> InlineKeyboardMarkup:
    """Меню выбора каналов"""
    builder = InlineKeyboardBuilder()
    for channel_id in SUPPLIER_CHANNELS:
        channel_name = get_channel_name(channel_id)
        builder.button(text=channel_name, callback_data=f"channel_select_{channel_id}")
    builder.button(text="🔙 Назад", callback_data="back_to_main")
    builder.adjust(1)
    return builder.as_markup()

def build_settings_menu() -> InlineKeyboardMarkup:
    """Меню настроек"""
    builder = InlineKeyboardBuilder()
    builder.button(text="⚙️ Настройки парсинга", callback_data="parse_settings")
    builder.button(text="💵 Правила ценообразования", callback_data="price_settings")
    builder.button(text="📁 Источники данных", callback_data="data_sources")
    builder.button(text="🔄 Обновить словари", callback_data="update_dict")
    builder.button(text="🔙 Назад", callback_data="back_to_main")
    builder.adjust(1)
    return builder.as_markup()

# =============================
# ОБРАБОТКА ГЛАВНОГО МЕНЮ
# =============================

@router.message(Command("start", "menu"))
async def start_command(message: Message):
    """Команда старта с выводом главного меню"""
    user = message.from_user
    welcome_text = (
        f"👋 Привет, {user.first_name}!\n"
        "🤖 Я бот для автоматизации работы с прайсами поставщиков.\n\n"
        "🔹 Используйте меню ниже для управления:"
    )
    await message.answer(welcome_text, reply_markup=build_main_menu())
    
    # Проверяем авторизацию Telethon
    if not DATA_STORAGE.get("telethon_client"):
        await init_telethon_client()

@router.callback_query(F.data == "back_to_main")
async def back_to_main_handler(callback: types.CallbackQuery):
    """Обработчик возврата в главное меню"""
    await callback.message.edit_text(
        "🏠 Главное меню:",
        reply_markup=build_main_menu()
    )
    await callback.answer()

# =============================
# ОБРАБОТКА КОМАНД МЕНЮ
# =============================

@router.callback_query(F.data == "get_messages")
async def get_messages_handler(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик команды загрузки сообщений"""
    if not SUPPLIER_CHANNELS:
        await callback.message.answer("❌ Не настроены каналы поставщиков. Используйте /settings")
        await callback.answer()
        return
    
    await callback.message.edit_text(
        "📥 Выберите канал поставщика:",
        reply_markup=build_channels_menu()
    )
    await callback.answer()

@router.callback_query(F.data == "process_data")
async def process_data_handler(callback: types.CallbackQuery):
    """Обработчик команды обработки данных"""
    user_id = callback.from_user.id
    
    # Проверяем наличие данных для текущего пользователя
    if "users" not in DATA_STORAGE or user_id not in DATA_STORAGE["users"]:
        await callback.message.answer(
            "❌ Сначала загрузите сообщения командой 'Загрузить сообщения'",
            reply_markup=build_back_button(MenuLevel.MAIN)
        )
        await callback.answer()
        return
    
    user_data = DATA_STORAGE["users"][user_id]
    
    # Проверка наличия необходимых данных
    if "last_message_ids" not in user_data or not user_data["last_message_ids"]:
        await callback.message.answer(
            "❌ Сначала загрузите сообщения командой 'Загрузить сообщения'",
            reply_markup=build_back_button(MenuLevel.MAIN)
        )
        await callback.answer()
        return
        
    if "last_channel" not in user_data:
        await callback.message.answer(
            "❌ Ошибка: канал не выбран. Начните заново",
            reply_markup=build_back_button(MenuLevel.MAIN)
        )
        await callback.answer()
        return
    
    # Вызываем функцию обработки данных
    await process_data_command(callback.message, user_id, user_data)
    await callback.answer()

@router.callback_query(F.data == "list_channels")
async def list_channels_handler(callback: types.CallbackQuery):
    """Обработчик команды списка каналов"""
    await list_channels_command(callback.message)
    await callback.answer()

@router.callback_query(F.data == "export_xlsx")
async def export_xlsx_handler(callback: types.CallbackQuery):
    """Обработчик команды экспорта в XLSX"""
    user_id = callback.from_user.id
    
    # Проверяем наличие данных для текущего пользователя
    if "users" not in DATA_STORAGE or user_id not in DATA_STORAGE["users"]:
        await callback.message.answer(
            "❌ Нет данных для экспорта. Сначала обработайте данные",
            reply_markup=build_back_button(MenuLevel.MAIN)
        )
        await callback.answer()
        return
        
    user_data = DATA_STORAGE["users"][user_id]
    
    if "processed_data" not in user_data or not user_data["processed_data"]:
        await callback.message.answer(
            "❌ Нет данных для экспорта. Сначала обработайте данные",
            reply_markup=build_back_button(MenuLevel.MAIN)
        )
        await callback.answer()
        return
    
    # Вызываем функцию экспорта
    await export_xlsx_command(callback.message, user_id, user_data)
    await callback.answer()

@router.callback_query(F.data == "settings")
async def settings_handler(callback: types.CallbackQuery):
    """Обработчик команды настроек"""
    await callback.message.edit_text(
        "⚙️ Настройки бота:",
        reply_markup=build_settings_menu()
    )
    await callback.answer()

@router.callback_query(F.data == "update_dict")
async def update_dict_handler(callback: types.CallbackQuery):
    """Обработчик обновления словарей"""
    await update_dict_command(callback.message)
    await callback.answer("🔄 Словари нормализации обновлены!")

# =============================
# ОБРАБОТЧИК ВОЗВРАТА НАЗАД
# =============================

@router.callback_query(F.data.startswith("back_to_"))
async def back_handler(callback: types.CallbackQuery):
    """Обработчик кнопки 'Назад'"""
    back_to = callback.data.replace("back_to_", "")
    
    if back_to == MenuLevel.MAIN:
        await callback.message.edit_text(
            "🏠 Главное меню:",
            reply_markup=build_main_menu()
        )
    elif back_to == MenuLevel.DATA_PROCESSING:
        await callback.message.edit_text(
            "📥 Выберите канал поставщика:",
            reply_markup=build_channels_menu()
        )
    elif back_to == MenuLevel.SETTINGS:
        await callback.message.edit_text(
            "⚙️ Настройки бота:",
            reply_markup=build_settings_menu()
        )
    
    await callback.answer()

# =============================
# ОБРАБОТКА ВЫБОРА КАНАЛА И СООБЩЕНИЙ
# =============================

@router.callback_query(F.data.startswith("channel_select_"))
async def process_channel_selection(callback: types.CallbackQuery, state: FSMContext):
    """Обработка выбора канала"""
    try:
        await callback.answer("⏳ Обработка...")
        
        channel_id = int(callback.data.replace("channel_select_", ""))
        await state.update_data(selected_channel=channel_id)
        
        channel_name = get_channel_name(channel_id)
        
        # Сохраняем ID сообщения для возможного возврата
        await state.update_data(prev_menu_message_id=callback.message.message_id)
        
        # Отправляем новое сообщение с инструкцией
        await callback.message.answer(
            f"✅ Выбран канал: <b>{channel_name}</b>\n"
            f"📥 Теперь введите ID сообщений через запятую или пробел:",
            reply_markup=build_back_button(MenuLevel.DATA_PROCESSING)
        )
        
        await state.set_state(Form.waiting_for_message_ids)
    except ValueError:
        await callback.message.answer("❌ Ошибка: неверный формат ID канала")
    except Exception as e:
        logger.error(f"Ошибка обработки выбора канала: {str(e)}", exc_info=True)
        await callback.message.answer("❌ Произошла ошибка при обработке выбора канала")
    finally:
        await callback.answer()

@router.message(Form.waiting_for_message_ids)
async def process_message_ids(message: Message, state: FSMContext):
    """Обработка введенных ID сообщений"""
    try:
        # Парсинг ID сообщений
        ids = [int(id_str.strip()) for id_str in re.split(r'[, \n]+', message.text) if id_str.strip().isdigit()]
        
        if not ids:
            await message.answer("❌ Не найдено валидных ID сообщений. Попробуйте снова.")
            return
        
        # Получение выбранного канала
        state_data = await state.get_data()
        selected_channel = state_data.get("selected_channel")
        
        if not selected_channel:
            await message.answer("❌ Ошибка: канал не выбран. Начните заново")
            await state.clear()
            return
        
        # Инициализация хранилища пользователя
        user_id = message.from_user.id
        if "users" not in DATA_STORAGE:
            DATA_STORAGE["users"] = {}
        if user_id not in DATA_STORAGE["users"]:
            DATA_STORAGE["users"][user_id] = {}
            
        # Сохраняем данные пользователя
        DATA_STORAGE["users"][user_id]["last_channel"] = selected_channel
        DATA_STORAGE["users"][user_id]["last_message_ids"] = ids
        
        channel_name = get_channel_name(selected_channel)
        await message.answer(
            f"✅ Получено {len(ids)} ID сообщений из канала {channel_name}. "
            f"Теперь вы можете обработать данные.",
            reply_markup=build_main_menu()  # Возврат в главное меню
        )
        logger.info(f"User {user_id} received {len(ids)} message IDs from channel {selected_channel}")
        
    except ValueError:
        await message.answer("❌ Ошибка формата. Вводите только числа, разделенные запятыми или пробелами.")
    finally:
        await state.clear()

# =============================
# ОБРАБОТКА ДАННЫХ
# =============================

def parse_supplier_message(text: str) -> Dict:
    """Парсинг сообщения поставщика с нормализацией данных"""
    # Удаление эмодзи и лишних символов (кроме специальных обозначений)
        
    if PARSING_SETTINGS.get("remove_emojis", True):
        cleaned = re.sub(r'[^\w\s.,;:!?@#$%^&*()\-+=\[\]{}<>/\\|♻️⌛️]', '', text)
    else:
        cleaned = text

    # Выявление категории товара
    category = "other"
    if "iPhone" in cleaned: category = "iPhone"
    elif "MacBook" in cleaned: category = "MacBook"
    elif "Apple Watch" in cleaned: category = "Apple Watch"
    elif "iPad" in cleaned: category = "iPad"
    elif "AirPods" in cleaned: category = "AirPods"
    elif "iMac" in cleaned: category = "iMac"
    elif "Samsung" in cleaned: category = "Samsung"
    elif "Xiaomi" in cleaned: category = "Xiaomi"
    elif "чехол" in cleaned or "кабель" in cleaned or "защитное стекло" in cleaned: 
        category = "Аксессуары"
    
    # Извлечение модели
    model_match = re.search(
        r'(iPhone \d+\w*|MacBook \w+|Watch \w+|iPad \w+|AirPods \w+|iMac \w+|Galaxy \w+|Redmi \w+|[А-Яа-я\w]+ \w+)', 
        cleaned, 
        re.IGNORECASE
    )
    model = model_match.group(0).strip() if model_match else "Неизвестная модель"
    
    # Извлечение цены
    price_match = re.search(r'(\d{4,6})\b', cleaned)
    price = int(price_match.group(1)) if price_match else 0
    
    # Извлечение наличия
    stock_match = re.search(r'(в наличии|на складе|доступно|количество):?\s*(\d+)', cleaned, re.IGNORECASE)
    stock = int(stock_match.group(2)) if stock_match else 0
    
    # Специфичная обработка для iPhone
    if category == "iPhone" and PARSING_SETTINGS.get("replace_flags", True):
        country_flags = {
            "🇺🇸": "US", "🇪🇺": "EU", "🇷🇺": "RU", "🇨🇳": "CH", 
            "🇬🇧": "UK", "🇯🇵": "JP", "🇰🇷": "KR", "🇭🇰": "HK"
        }
        for flag, code in country_flags.items():
            cleaned = cleaned.replace(flag, code)

                # Извлечение региона
        region_match = re.search(r'\b(US|EU|RU|CH|UK|JP|KR|HK|TW|SG)\b', cleaned)
        region = region_match.group(0) if region_match else "XX"
        model += f" {region}"
        
    # Обработка типов SIM, если включено в настройках
    if category == "iPhone" and PARSING_SETTINGS.get("process_sim_types", True):
        cleaned = re.sub(r'(2SIM|DualSIM)', '2SIM', cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r'(eSIM|SIM\+eSIM)', 'eSIM', cleaned, flags=re.IGNORECASE)

    
    # Обработка обозначений поставки
    delivery_icons = {"♻️": "Рефабрикат", "⌛️": "Ожидается"}
    for icon, desc in delivery_icons.items():
        if icon in text:
            model += f" ({desc})"
    
    return {
        "category": category,
        "model": model,
        "price": price,
        "stock": stock,
        "original_text": text,
        "cleaned_text": cleaned
    }

def normalize_data(messages: List[Dict]) -> List[Dict]:
    """Нормализация и объединение данных"""
    normalized = []
    seen_models = set()
    
    for msg in messages:
        if not msg: continue
        
        model_key = (msg["category"], msg["model"])
        if model_key in seen_models:
            # Объединение данных по существующей модели
            for item in normalized:
                if item["category"] == msg["category"] and item["model"] == msg["model"]:
                    item["stock"] += msg["stock"]
                    # Сохраняем оригинальные тексты для отладки
                    item["original_text"] += f"\n\n---\n\n{msg['original_text']}"
                    break
        else:
            seen_models.add(model_key)
            normalized.append(msg)
    
    # Сортировка: сначала Apple продукты, потом другие
    apple_categories = {"iPhone", "MacBook", "Apple Watch", "iPad", "AirPods", "iMac"}
    apple_items = [item for item in normalized if item["category"] in apple_categories]
    other_items = [item for item in normalized if item["category"] not in apple_categories]
    
    # Специфичная сортировка для Apple продуктов
    def apple_sort_key(x):
        model = x["model"]
        # Приоритет по типу продукта
        type_order = {"iPhone": 0, "MacBook": 1, "iPad": 2, "Apple Watch": 3, "AirPods": 4, "iMac": 5}
        type_priority = type_order.get(x["category"], 99)
        
        # Сортировка iPhone по номеру модели и объему памяти
        if "iPhone" in model:
            model_num = int(re.search(r'iPhone (\d+)', model).group(1)) if re.search(r'iPhone (\d+)', model) else 0
            storage_match = re.search(r'(\d+)GB', model)
            storage = int(storage_match.group(1)) if storage_match else 0
            return (0, model_num, storage, -x["price"])
        
        return (type_priority, model)
    
    apple_items.sort(key=apple_sort_key)
    
    # Сортировка других товаров по категории и модели
    other_items.sort(key=lambda x: (x["category"], x["model"]))
    
    return apple_items + other_items

def calculate_prices(data: List[Dict]) -> List[Dict]:
    """Расчет цен с применением правил наценки"""

    global PRICE_SETTINGS

    processed = []
    max_margin_product = {"margin": -1, "index": -1}
    
    for idx, item in enumerate(data):
        category = item["category"]
        base_price = item["price"]
        
        # Пропуск товаров с нулевой ценой
        if base_price <= 0:
            item["final_price"] = 0
            item["margin"] = 0
            processed.append(item)
            continue
        
        # Применение правил ценообразования
         # Применение правил ценообразования
        rules = PRICE_SETTINGS.get(category, PRICE_SETTINGS.get("default", {"margin": 0.1}))
        
        if "base_margin" in rules:
            final_price = base_price * (1 + rules["base_margin"])
            # Дополнительная наценка для премиум моделей
            if "premium_bonus" in rules and ("Pro" in item["model"] or "Max" in item["model"]):
                final_price *= (1 + rules["premium_bonus"])
        elif "fixed_margin" in rules:
            final_price = base_price + rules["fixed_margin"]
        else:
            final_price = base_price * 1.1  # Дефолтная наценка 10%
        
        # Округление до сотен
        final_price = round(final_price / 100) * 100
        margin_value = final_price - base_price
        
        item["final_price"] = final_price
        item["margin"] = margin_value
        
        # Поиск товара с максимальной маржой в категории Apple
        if category in PRICE_RULES and margin_value > max_margin_product["margin"]:
            max_margin_product = {"margin": margin_value, "index": idx}
    
    # Добавление смайлика к товару с максимальной маржой
    if max_margin_product["index"] >= 0:
        data[max_margin_product["index"]]["model"] += " 😊"
    
    return data

# =============================
# РАБОТА С ФАЙЛАМИ
# =============================

def generate_xlsx(data: List[Dict]) -> str:
    """Генерация XLSX файла с форматированием"""
    filename = f"price_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"
    
    # Заголовки
    headers = ["Категория", "Модель", "Цена поставщика", "Цена на сайте", "Наличие", "Маржа"]
    ws.append(headers)
    
    # Форматирование заголовков
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Данные
    for item in data:
        # Форматирование цены на сайте (пробел после тысяч)
        site_price = item['final_price']
        if site_price > 0:
            site_price_str = f"{site_price // 1000} {str(site_price)[-3:]}"
        else:
            site_price_str = "0"
        
        row = [
            item["category"],
            item["model"],
            item["price"],
            site_price_str,
            item["stock"],
            item.get("margin", 0)
        ]
        ws.append(row)
    
    # Форматирование столбца с ценами
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value and cell.value != "0":
                cell.font = Font(bold=True, color="FF0000")
    
    # Автонастройка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    wb.save(filename)
    return filename

async def delete_file_after_delay(filename: str, delay: int = 120):
    """Удаление файла после задержки"""
    await asyncio.sleep(delay)
    try:
        if os.path.exists(filename):
            os.remove(filename)
            logger.info(f"Файл {filename} успешно удален")
    except Exception as e:
        logger.error(f"Ошибка удаления файла {filename}: {str(e)}")

# =============================
# ОСНОВНЫЕ ФУНКЦИИ БОТА
# =============================

@router.message(Command("code"))
async def process_verification_code(message: Message):
    """Обработка кода подтверждения"""
    code = message.text.split(maxsplit=1)[1] if len(message.text.split()) > 1 else None
    
    if not code or not code.isdigit():
        await message.answer("❌ Неверный формат кода. Используйте: /code <ваш_код>")
        return
    
    client = TelegramClient(SESSION_NAME, API_ID, API_HASH)
    
    try:
        await client.start(phone=lambda: PHONE_NUMBER, code=code)
        
        if await client.is_user_authorized():
            DATA_STORAGE["telethon_client"] = client
            await message.answer("✅ Авторизация прошла успешно! Бот готов к работе.")
            logger.info("Telethon client authorized successfully")
            # Обновляем названия каналов
            await refresh_channel_names()
        else:
            await message.answer("❌ Не удалось авторизоваться. Попробуйте снова.")
    except Exception as e:
        await message.answer(f"❌ Ошибка авторизации: {str(e)}")
        logger.error(f"Authorization error: {str(e)}")

@router.message(F.forward_from_chat)
async def handle_forwarded_message(message: Message):
    """Автоматическая обработка пересланных сообщений"""
    chat = message.forward_from_chat
    msg_id = message.forward_from_message_id
    
    # Для каналов chat.id будет отрицательным числом
    channel_id = chat.id
    
    # Проверка доступа к каналу
    if channel_id not in SUPPLIER_CHANNELS:
        await message.answer(
            f"⚠️ Канал {get_channel_name(channel_id)} не в списке разрешенных. "
            f"Добавьте его в настройки или используйте команду 'Загрузить сообщения'."
        )
        return
    
    # Сохраняем для обработки
    user_id = message.from_user.id
    if "users" not in DATA_STORAGE:
        DATA_STORAGE["users"] = {}
    if user_id not in DATA_STORAGE["users"]:
        DATA_STORAGE["users"][user_id] = {}
        
    DATA_STORAGE["users"][user_id]["last_channel"] = channel_id
    DATA_STORAGE["users"][user_id]["last_message_ids"] = [msg_id]
    
    # Предлагаем обработать
    keyboard = InlineKeyboardBuilder()
    keyboard.button(text="🔄 Обработать сообщение", callback_data="process_now")
    
    await message.answer(
        f"📥 Сообщение из {get_channel_name(channel_id)} получено!\n"
        f"• ID сообщения: {msg_id}\n\n"
        f"Хотите обработать его сейчас?",
        reply_markup=keyboard.as_markup()
    )

@router.callback_query(F.data == "process_now")
async def process_now_callback(callback: types.CallbackQuery):
    """Обработка по нажатию кнопки"""
    await callback.message.edit_reply_markup()  # Убираем кнопки
    
    user_id = callback.from_user.id
    if "users" not in DATA_STORAGE or user_id not in DATA_STORAGE["users"]:
        await callback.message.answer("❌ Ошибка: данные не найдены")
        return
        
    user_data = DATA_STORAGE["users"][user_id]
    await process_data_command(callback.message, user_id, user_data)
    await callback.answer()

async def list_channels_command(message: Message):
    """Функция списка каналов"""
    if not SUPPLIER_CHANNELS:
        await message.answer("ℹ️ Список каналов поставщиков не настроен")
        return
        
    # Обновляем названия каналов
    await refresh_channel_names()
    
    channel_list = "\n".join([
        f"• {get_channel_name(channel_id)} (ID: `{channel_id}`)" 
        for channel_id in SUPPLIER_CHANNELS
    ])
    
    await message.answer(
        f"📋 <b>Доступные каналы поставщиков:</b>\n\n{channel_list}\n\n"
        f"ℹ️ Для добавления новых каналов измените переменную окружения SUPPLIER_CHANNELS",
        reply_markup=build_back_button(MenuLevel.MAIN)
    )

async def update_dict_command(message: Message):
    """Функция обновления словарей"""
    await message.answer("🔄 Словари нормализации обновлены!")
    logger.info("Normalization dictionaries updated by user")

async def process_data_command(message: Message, user_id: int, user_data: dict):
    """Функция обработки данных"""
    # Проверка наличия необходимых данных
    if "last_message_ids" not in user_data or not user_data["last_message_ids"]:
        await message.answer("❌ Сначала загрузите сообщения командой 'Загрузить сообщения'")
        return
        
    if "last_channel" not in user_data:
        await message.answer("❌ Ошибка: канал не выбран. Начните заново")
        return
        
    client = DATA_STORAGE.get("telethon_client")
    if not client:
        await message.answer("❌ Telethon клиент не инициализирован. Попробуйте позже.")
        return
        
    message_ids = user_data["last_message_ids"]
    channel_id = user_data["last_channel"]
    channel_name = get_channel_name(channel_id)
    
    await message.answer(f"⏳ Загружаю {len(message_ids)} сообщений из {channel_name}...")
    
    try:
        # Получение сообщений через Telethon
        messages = await fetch_messages(channel_id, message_ids)
        
        if not messages:
            await message.answer("❌ Не удалось загрузить сообщения. Проверьте ID и доступ к каналу.")
            return
            
        # Парсинг и обработка данных
        await message.answer("🔄 Парсинг и нормализация данных...")
        parsed = [parse_supplier_message(text) for text in messages.values()]
        normalized = normalize_data(parsed)
        
        await message.answer("🧮 Расчет цен с наценками...")
        processed = calculate_prices(normalized)
        
        # Сохранение результатов
        user_data["processed_data"] = processed
        user_data["raw_messages"] = messages
        
        # Формирование отчета
        report = "✅ Данные успешно обработаны!\n\n"
        report += f"🔹 Канал: {channel_name}\n"
        report += f"🔹 Сообщений: {len(messages)}\n"
        report += f"🔹 Товаров: {len(processed)}\n"
        report += f"🔹 Категории: {', '.join(set(item['category'] for item in processed))}\n\n"
        report += "ℹ️ Теперь вы можете экспортировать данные в XLSX"
        
        await message.answer(report)
        logger.info(f"User {user_id} processed {len(processed)} items from {len(messages)} messages")
        
    except Exception as e:
        await message.answer(f"❌ Ошибка при обработке данных: {str(e)}")
        logger.error(f"Data processing error for user {user_id}: {str(e)}", exc_info=True)

async def export_xlsx_command(message: Message, user_id: int, user_data: dict):
    """Генерация и отправка XLSX файла"""
    if "processed_data" not in user_data or not user_data["processed_data"]:
        await message.answer("❌ Нет данных для экспорта. Сначала обработайте данные")
        return
    
    try:
        await message.answer("⏳ Формирую XLSX файл...")
        filename = generate_xlsx(user_data["processed_data"])
        
        # Используем FSInputFile для файлов в файловой системе
        file = FSInputFile(filename)
        
        await message.answer_document(
            document=file,
            caption=f"📊 Прайс-лист от {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )
        logger.info(f"User {user_id} exported XLSX file: {filename}")
        
        # Удаляем временный файл через 2 минуты
        asyncio.create_task(delete_file_after_delay(filename, 120))
        
    except Exception as e:
        await message.answer(f"❌ Ошибка при создании файла: {str(e)}")
        logger.error(f"XLSX export error for user {user_id}: {str(e)}", exc_info=True)

# =============================
# ОБРАБОТЧИКИ НАСТРОЕК
# =============================

@router.callback_query(F.data == "parse_settings")
async def parse_settings_handler(callback: types.CallbackQuery):
    """Обработчик настроек парсинга"""
    builder = InlineKeyboardBuilder()
    
    # Кнопки для каждого параметра
    for key, value in PARSING_SETTINGS.items():
        if isinstance(value, bool):
            builder.button(
                text=f"{key}: {'✅' if value else '❌'}",
                callback_data=f"toggle_parse_{key}"
            )
    
    builder.button(text="✏️ Редактировать иконки", callback_data="edit_delivery_icons")
    builder.button(text="🔙 Назад", callback_data="back_to_settings")
    builder.adjust(1)
    
    await callback.message.edit_text(
        "⚙️ <b>Настройки парсинга:</b>\n\n"
        "• Удаление эмодзи\n"
        "• Замена флагов регионов\n"
        "• Обработка типов SIM\n"
        "• Обработка иконок поставки",
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@router.callback_query(F.data.startswith("toggle_parse_"))
async def toggle_parse_setting(callback: types.CallbackQuery):
    """Переключение параметра парсинга"""
    key = callback.data.replace("toggle_parse_", "")
    
    if key in PARSING_SETTINGS:
        PARSING_SETTINGS[key] = not PARSING_SETTINGS[key]
        save_parsing_settings()
        await parse_settings_handler(callback)  # Обновляем меню
    await callback.answer()

@router.callback_query(F.data == "edit_delivery_icons")
async def edit_delivery_icons(callback: types.CallbackQuery, state: FSMContext):
    """Редактирование иконок поставки"""
    icons_text = "\n".join([f"{icon} - {desc}" for icon, desc in PARSING_SETTINGS["delivery_icons"].items()])
    
    await callback.message.answer(
        f"📝 Текущие иконки поставки:\n{icons_text}\n\n"
        "Введите новые иконки в формате:\n"
        "♻️=Рефабрикат\n"
        "⌛️=Ожидается\n"
        "🚚=Доставка",
        reply_markup=build_back_button(MenuLevel.SETTINGS)
    )
    await state.set_state(Form.waiting_for_parsing_settings)
    await callback.answer()

@router.message(Form.waiting_for_parsing_settings)
async def process_parsing_settings(message: Message, state: FSMContext):
    """Обработка новых настроек парсинга"""
    try:
        new_icons = {}
        for line in message.text.split('\n'):
            if '=' in line:
                icon, desc = line.split('=', 1)
                new_icons[icon.strip()] = desc.strip()
        
        PARSING_SETTINGS["delivery_icons"] = new_icons
        save_parsing_settings()
        await message.answer("✅ Иконки поставки успешно обновлены!")
    except Exception as e:
        await message.answer(f"❌ Ошибка обработки: {str(e)}")
    finally:
        await state.clear()

@router.callback_query(F.data == "price_settings")
async def price_settings_handler(callback: types.CallbackQuery):
    """Обработчик правил ценообразования"""
    builder = InlineKeyboardBuilder()
    
    # Кнопки для каждой категории
    for category in PRICE_SETTINGS.keys():
        if category != "default":
            builder.button(text=category, callback_data=f"price_category_{category}")
    
    builder.button(text="➕ Добавить категорию", callback_data="add_price_category")
    builder.button(text="🔙 Назад", callback_data="back_to_settings")
    builder.adjust(2)
    
    await callback.message.edit_text(
        "💵 <b>Правила ценообразования:</b>\n\n"
        "Выберите категорию для редактирования:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@router.callback_query(F.data.startswith("price_category_"))
async def edit_price_category(callback: types.CallbackQuery, state: FSMContext):
    """Редактирование правил для категории"""
    category = callback.data.replace("price_category_", "")
    
    # Сохраняем категорию в состоянии
    await state.update_data(price_category=category)
    
    # Получаем текущие правила
    rules = PRICE_SETTINGS.get(category, {})
    rules_text = "\n".join([f"{key}: {value}" for key, value in rules.items()])
    
    await callback.message.answer(
        f"📝 Правила для <b>{category}</b>:\n{rules_text}\n\n"
        "Введите новые правила в формате:\n"
        "base_margin=0.15\n"
        "premium_bonus=0.1\n"
        "fixed_margin=5000",
        reply_markup=build_back_button(MenuLevel.SETTINGS)
    )
    await state.set_state(Form.waiting_for_price_value)
    await callback.answer()

@router.message(Form.waiting_for_price_value)
async def process_price_settings(message: Message, state: FSMContext):
    """Обработка новых правил ценообразования"""
    try:
        state_data = await state.get_data()
        category = state_data.get("price_category")
        
        if not category:
            await message.answer("❌ Категория не выбрана")
            await state.clear()
            return
        
        new_rules = {}
        for line in message.text.split('\n'):
            if '=' in line:
                key, value = line.split('=', 1)
                try:
                    # Пробуем преобразовать в число
                    new_rules[key.strip()] = float(value.strip())
                except ValueError:
                    new_rules[key.strip()] = value.strip()
        
        PRICE_SETTINGS[category] = new_rules
        save_price_settings()
        await message.answer(f"✅ Правила для <b>{category}</b> успешно обновлены!")
    except Exception as e:
        await message.answer(f"❌ Ошибка обработки: {str(e)}")
    finally:
        await state.clear()

@router.callback_query(F.data == "add_price_category")
async def add_price_category(callback: types.CallbackQuery, state: FSMContext):
    """Добавление новой категории"""
    await callback.message.answer(
        "📝 Введите название новой категории:",
        reply_markup=build_back_button(MenuLevel.SETTINGS)
    )
    await state.set_state(Form.waiting_for_price_category)
    await callback.answer()

@router.message(Form.waiting_for_price_category)
async def process_new_category(message: Message, state: FSMContext):
    """Обработка новой категории"""
    try:
        category = message.text.strip()
        if not category:
            await message.answer("❌ Название категории не может быть пустым")
            return
        
        if category not in PRICE_SETTINGS:
            PRICE_SETTINGS[category] = {"base_margin": 0.1}
            save_price_settings()
            await message.answer(f"✅ Категория <b>{category}</b> успешно добавлена!")
        else:
            await message.answer("ℹ️ Такая категория уже существует")
    except Exception as e:
        await message.answer(f"❌ Ошибка обработки: {str(e)}")
    finally:
        await state.clear()

@router.callback_query(F.data == "data_sources")
async def data_sources_handler(callback: types.CallbackQuery):
    """Обработчик источников данных"""
    builder = InlineKeyboardBuilder()
    
    # Кнопки для каждого канала
    for channel_id in DATA_SOURCES:
        channel_name = get_channel_name(channel_id)
        builder.button(text=channel_name, callback_data=f"view_channel_{channel_id}")
    
    builder.button(text="➕ Добавить канал", callback_data="add_data_source")
    builder.button(text="🔙 Назад", callback_data="back_to_settings")
    builder.adjust(1)
    
    await callback.message.edit_text(
        "📁 <b>Источники данных:</b>\n\n"
        "Выберите канал для управления:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@router.callback_query(F.data.startswith("view_channel_"))
async def view_channel_details(callback: types.CallbackQuery):
    """Просмотр деталей канала"""
    channel_id = int(callback.data.replace("view_channel_", ""))
    channel_name = get_channel_name(channel_id)
    
    builder = InlineKeyboardBuilder()
    builder.button(text="🗑️ Удалить", callback_data=f"remove_channel_{channel_id}")
    builder.button(text="🔙 Назад", callback_data="back_to_data_sources")
    builder.adjust(1)
    
    await callback.message.edit_text(
        f"📋 <b>Информация о канале:</b>\n\n"
        f"• ID: <code>{channel_id}</code>\n"
        f"• Название: {channel_name}",
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@router.callback_query(F.data.startswith("remove_channel_"))
async def remove_channel(callback: types.CallbackQuery):
    """Удаление канала из источников"""
    channel_id = int(callback.data.replace("remove_channel_", ""))
    
    if channel_id in DATA_SOURCES:
        DATA_SOURCES.remove(channel_id)
        save_data_sources()
        await callback.answer(f"✅ Канал удален из источников")
        await data_sources_handler(callback)  # Возвращаемся к списку
    else:
        await callback.answer("ℹ️ Канал не найден в списке")

@router.callback_query(F.data == "add_data_source")
async def add_data_source(callback: types.CallbackQuery, state: FSMContext):
    """Добавление нового источника данных"""
    await callback.message.answer(
        "📝 Введите ID нового канала (например, -100123456789):",
        reply_markup=build_back_button(MenuLevel.SETTINGS)
    )
    await state.set_state(Form.waiting_for_new_channel)
    await callback.answer()

@router.message(Form.waiting_for_new_channel)
async def process_new_channel(message: Message, state: FSMContext):
    """Обработка нового канала"""
    try:
        channel_id = int(message.text.strip())
        if channel_id not in DATA_SOURCES:
            DATA_SOURCES.append(channel_id)
            save_data_sources()
            
            # Обновляем кэш названий
            await refresh_channel_names()
            
            await message.answer(f"✅ Канал <code>{channel_id}</code> успешно добавлен!")
        else:
            await message.answer("ℹ️ Этот канал уже есть в списке источников")
    except ValueError:
        await message.answer("❌ Неверный формат ID. Введите числовой ID канала")
    except Exception as e:
        await message.answer(f"❌ Ошибка обработки: {str(e)}")
    finally:
        await state.clear()

@router.callback_query(F.data == "back_to_data_sources")
async def back_to_data_sources(callback: types.CallbackQuery):
    """Возврат к списку источников данных"""
    await data_sources_handler(callback)

async def refresh_channel_names():
    """Обновление кэша названий каналов с учетом текущих источников"""
    client = DATA_STORAGE.get("telethon_client")
    if not client:
        return
    
    for channel_id in DATA_SOURCES:  # Используем глобальный список источников
        try:
            entity = await client.get_entity(channel_id)
            DATA_STORAGE["channel_names"][channel_id] = entity.title
        except Exception as e:
            logger.error(f"Ошибка получения названия канала {channel_id}: {str(e)}")
            DATA_STORAGE["channel_names"][channel_id] = f"Канал {channel_id}"

# =============================
# ЗАПУСК И ОСТАНОВКА БОТА
# =============================

async def on_startup(dispatcher, bot):
    """Действия при запуске бота"""
    logger.info("Запуск бота...")
    await init_telethon_client()
    
    # Обновляем названия каналов
    await refresh_channel_names()
    
    # Планировщик для периодических задач
    scheduler = AsyncIOScheduler()
    scheduler.start()
    
    logger.info("Бот успешно запущен")

async def on_shutdown(dispatcher, bot):
    """Действия при остановке бота"""
    logger.info("Остановка бота...")
    await shutdown_telethon_client()
    logger.info("Бот успешно остановлен")

async def main():
    """Основная функция запуска"""
    logger.info("Starting price bot...")
    
    # Регистрация обработчиков запуска/остановки
    dp.startup.register(on_startup)
    dp.shutdown.register(on_shutdown)
    
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())